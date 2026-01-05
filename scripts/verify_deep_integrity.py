import json
import openpyxl
from supabase import create_client, Client

# Setup
SUPABASE_URL = 'https://pevstuyzlewvjidjkmea.supabase.co'
SUPABASE_KEY = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InBldnN0dXl6bGV3dmppZGprbWVhIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NjU3NzczNTYsImV4cCI6MjA4MTM1MzM1Nn0.myYaq8rshNyB2aGTas2f1IzsQVv_rihOGL2v8EPl-x0'
JSON_PATH = r"c:\Antigravity - RVM Designações\backup_rvm_2026-01-05.json"
XLSX_PATH = r"c:\Antigravity - RVM Designações\backup_rvm_2026-01-05.xlsx"

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

TABLES_TO_CHECK = [
    'workbook_parts',
    'special_events',
    'workbook_batches',
    'publishers',
    'local_needs_preassignments',
    'extraction_history'
]

def get_db_schema_sample(table):
    """Fetches one row from DB to infer schema"""
    res = supabase.table(table).select('*').limit(1).execute()
    if res.data and len(res.data) > 0:
        return res.data[0]
    return None

def normalize_value(v):
    if v is None: return "null"
    if isinstance(v, bool): return "true" if v else "false"
    return str(v).strip()

def compare_objects(db_obj, file_obj, table_name, file_type="JSON"):
    # 1. Compare Keys
    db_keys = set(db_obj.keys())
    file_keys = set(file_obj.keys())
    
    # Filter out internal openpyxl stuff or ignore some keys if needed
    # (In XL, keys are columns headers)
    
    print(f"      Comparing ID {db_obj.get('id')} ({file_type})...")
    
    diffs = []
    for k in db_keys:
        if k in file_keys:
            db_val = db_obj[k]
            file_val = file_obj[k]

            # Special handling for JSON strings in Excel
            if file_type == "XLSX":
                if k in ['configuration', 'details', 'data'] and isinstance(file_val, str) and file_val.startswith('{'):
                    try:
                        file_val = json.loads(file_val)
                    except:
                        pass
            
            # Helper for object comparison (dicts)
            if isinstance(db_val, dict) and isinstance(file_val, dict):
                 if json.dumps(db_val, sort_keys=True) != json.dumps(file_val, sort_keys=True):
                      diffs.append(f"{k}: Object mismatch")
                 continue

            norm_db = normalize_value(db_val)
            norm_file = normalize_value(file_val)
            
            # Date precision hack
            if 'created_at' in k or 'date' in k:
                if norm_db[:19] == norm_file[:19]:
                    continue
            
            if norm_db != norm_file:
                diffs.append(f"{k}: DB='{norm_db}' vs FILE='{norm_file}'")
    
    if not diffs:
        print(f"      ✅ OK")
    else:
        print(f"      ❌ DIFERENÇAS: {diffs}")

def verify_deep():
    print("--- DEEP INTEGRITY CHECK (JSON & XLSX) ---")
    
    # Load Files
    try:
        with open(JSON_PATH, 'r', encoding='utf-8') as f:
            json_backup = json.load(f)
        xlsx_wb = openpyxl.load_workbook(XLSX_PATH)
    except Exception as e:
        print(f"Erro carregando arquivos: {e}")
        return

    json_tables = json_backup.get('tables', {})

    for table in TABLES_TO_CHECK:
        print(f"\n[{table}]")
        try:
            # Get DB Sample
            print(f"  > Fetching DB sample...")
            db_sample = get_db_schema_sample(table)
            if not db_sample:
                print(f"      ⚠️ Tabela vazia no banco ou erro de acesso.")
                continue
            
            target_id = db_sample.get('id')
            print(f"      Using ID: {target_id}")

            # 1. Check JSON
            print(f"  > Validating JSON...")
            json_rows = json_tables.get(table, {}).get('data', [])
            json_match = next((item for item in json_rows if item.get('id') == target_id), None)
            if json_match:
                compare_objects(db_sample, json_match, table, "JSON")
            else:
                print(f"      ❌ ID não encontrado no JSON")

            # 2. Check XLSX
            print(f"  > Validating XLSX...")
            if table in xlsx_wb.sheetnames:
                ws = xlsx_wb[table]
                # Leitura mais robusta de headers
                headers = [cell.value for cell in ws[1]]
                
                xlsx_match = None
                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_dict = dict(zip(headers, row))
                    if row_dict.get('id') == target_id:
                        xlsx_match = row_dict
                        break
                
                if xlsx_match:
                    compare_objects(db_sample, xlsx_match, table, "XLSX")
                else:
                     print(f"      ❌ ID não encontrado no XLSX")
            else:
                print(f"      ⚠️ Aba '{table}' não encontrada no XLSX")

        except Exception as e:
            print(f"  ❌ Erro ao verificar tabela '{table}': {e}")

if __name__ == "__main__":
    verify_deep()

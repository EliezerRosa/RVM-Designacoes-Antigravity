import json
import openpyxl
from supabase import create_client, Client
import os

# Configuração
SUPABASE_URL = 'https://pevstuyzlewvjidjkmea.supabase.co'
SUPABASE_KEY = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InBldnN0dXl6bGV3dmppZGprbWVhIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NjU3NzczNTYsImV4cCI6MjA4MTM1MzM1Nn0.myYaq8rshNyB2aGTas2f1IzsQVv_rihOGL2v8EPl-x0'

JSON_PATH = r"c:\Antigravity - RVM Designações\backup_rvm_2026-01-05.json"
XLSX_PATH = r"c:\Antigravity - RVM Designações\backup_rvm_2026-01-05.xlsx"

EXPECTED_TABLES = [
    'publishers',
    'workbook_parts',
    'workbook_batches',
    'special_events',
    'extraction_history',
    'local_needs_preassignments'
]

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

def verify_all_tables():
    print(f"\n--- Verificando Todas as Tabelas ---")
    
    # 1. Ler JSON para descobrir tabelas
    try:
        with open(JSON_PATH, 'r', encoding='utf-8') as f:
            json_data = json.load(f)
        
        tables_data = json_data.get('tables', {})
        if not tables_data:
            print("⚠️ Nenhuma tabela encontrada na chave 'tables' do JSON.")
            return

        found_keys = list(tables_data.keys())
        print(f"Tabelas encontradas no JSON: {found_keys}")
        
        # Verificar faltantes
        missing = [t for t in EXPECTED_TABLES if t not in found_keys]
        if missing:
             print(f"⚠️ ATENÇÃO: Tabelas esperadas NÃO encontradas no JSON: {missing}")
        else:
             print("✅ Todas as tabelas críticas estão presentes no JSON.")

    except Exception as e:
        print(f"❌ Erro ao ler JSON: {e}")
        return

    # 2. Ler XLSX para descobrir abas
    xlsx_counts = {}
    try:
        wb = openpyxl.load_workbook(XLSX_PATH)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Assumindo 1 linha de header
            count = max(0, ws.max_row - 1)
            xlsx_counts[sheet_name] = count
    except Exception as e:
        print(f"❌ Erro ao ler XLSX: {e}")

    # 3. Comparar cada tabela
    print("\n" + "="*80)
    print(f"{'TABELA':<30} | {'DB':<10} | {'JSON':<10} | {'XLSX':<10} | {'STATUS':<15}")
    print("-" * 80)

    for table_name in tables_data.keys():
        # Contagem JSON
        json_rows = tables_data[table_name].get('data', [])
        json_count = len(json_rows)
        
        # Contagem DB
        try:
            db_res = supabase.table(table_name).select('*', count='exact', head=True).execute()
            db_count = db_res.count if db_res.count is not None else -1
        except Exception as e:
            db_count = "N/A" # Tabela pode não existir no DB ou erro de permissão
        
        # Contagem XLSX (tenta match exato ou case-insensitive)
        xlsx_count = xlsx_counts.get(table_name, "N/A")
        if xlsx_count == "N/A":
             # Tentar achar ignorando case
             for sheet in xlsx_counts:
                 if sheet.lower() == table_name.lower():
                     xlsx_count = xlsx_counts[sheet]
                     break

        # Status
        status = "✅ OK"
        if db_count != "N/A" and db_count != json_count:
            status = "❌ JSON DIFF"
        if xlsx_count != "N/A" and isinstance(xlsx_count, int) and isinstance(db_count, int):
             if db_count != xlsx_count:
                 status = "❌ XLSX DIFF"
        
        print(f"{table_name:<30} | {str(db_count):<10} | {str(json_count):<10} | {str(xlsx_count):<10} | {status:<15}")

if __name__ == "__main__":
    print("INICIANDO VERIFICAÇÃO COMPLETA DE BACKUP...")
    verify_all_tables()

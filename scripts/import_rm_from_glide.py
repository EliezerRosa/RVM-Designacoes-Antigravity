#!/usr/bin/env python3
"""
import_rm_from_glide.py — Bootstrap one-shot do schema rm.* a partir do export Glide.

Fontes:
  --ods  docs/RM Desacoplado/9fe36d.Relatório Mensal v03 (New).ods  (padrão)
           aba Congregação    -> rm.congregations
           aba Grupos         -> rm.field_groups (líderes resolvidos em 2ª passada)
           aba PublicadorReal -> rm.publishers
         Fallback se --ods ausente/inexistente: CSVs em --dir
  --reports  OneDrive/Relatórios Glide.xlsx  -> rm.monthly_reports  (padrão canônico)

Conexão: Postgres DIRETO (psycopg2), bypassa PostgREST e RLS.
Env obrigatória:
  RM_DATABASE_URL = postgresql://postgres:<pwd>@<host>:5432/postgres

Dependências:
  pip install psycopg2-binary openpyxl odfpy pandas

Uso mínimo (usa defaults):
  python import_rm_from_glide.py
Uso explícito:
  python import_rm_from_glide.py --ods "docs/RM Desacoplado/9fe36d.Relatório Mensal v03 (New).ods"
"""
from __future__ import annotations

import argparse
import csv
import os
import sys
import unicodedata
from pathlib import Path
from typing import Any

try:
    import psycopg2
    import psycopg2.extras
except ImportError:
    sys.exit("Falta psycopg2. Rode no venv: pip install psycopg2-binary")

try:
    import pandas as pd  # noqa: F401  (usado em read_ods_sheet)
except ImportError:
    sys.exit("Falta pandas. Rode no venv: pip install pandas odfpy")


# ----------------------------------------------------------------------------
# Utilitários de normalização (headers Glide têm acentos e variações)
# ----------------------------------------------------------------------------
def norm_key(s: str) -> str:
    """casefold + remove acentos + só alfanumérico, para casar headers robustamente."""
    s = unicodedata.normalize("NFD", s or "")
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return "".join(c for c in s.lower() if c.isalnum())


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    # utf-8-sig lida com BOM; Glide exporta UTF-8.
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.DictReader(fh)
        rows = list(reader)
    return rows


def read_ods_sheet(ods_path: Path, sheet_name: str) -> list[dict[str, str]]:
    """Lê uma aba do .ods e retorna lista de dicts (mesmo formato de read_csv_rows)."""
    import pandas as pd
    df = pd.read_excel(ods_path, engine="odf", sheet_name=sheet_name, dtype=str)
    df = df.where(df.notna(), other="")  # NaN -> string vazia
    return df.to_dict(orient="records")


def col(row: dict[str, str], *candidates: str) -> str | None:
    """Busca valor por qualquer um dos nomes de coluna candidatos (normalizados)."""
    index = {norm_key(k): v for k, v in row.items()}
    for c in candidates:
        v = index.get(norm_key(c))
        if v is not None and v.strip() != "":
            return v.strip()
    return None


def to_bool(v: str | None) -> bool:
    return (v or "").strip().lower() in ("true", "sim", "1", "t", "yes")


def to_date(v: str | None) -> str | None:
    """Converte datas Glide (dd/mm/yyyy ou mm/dd/yyyy) para ISO; None se vazio."""
    if not v or not v.strip():
        return None
    raw = v.strip().split(" ")[0].split(",")[0]
    for sep in ("/", "-"):
        if sep in raw:
            parts = raw.split(sep)
            if len(parts) == 3:
                a, b, c = parts
                if len(a) == 4:  # yyyy-mm-dd
                    return f"{a}-{int(b):02d}-{int(c):02d}"
                # dd/mm/yyyy (padrão BR das CSVs)
                try:
                    d, m, y = int(a), int(b), int(c)
                    if m > 12:  # provavelmente mm/dd
                        d, m = m, d
                    return f"{y:04d}-{m:02d}-{d:02d}"
                except ValueError:
                    return None
    return None


def gender(v: str | None) -> str | None:
    if not v:
        return None
    u = v.strip().upper()
    return "M" if u.startswith("M") else ("F" if u.startswith("F") else None)


def status_norm(v: str | None) -> str | None:
    if not v:
        return None
    u = norm_key(v)
    for st in ("ATIVO", "IRREGULAR", "QUASEINATIVO", "INATIVO"):
        if norm_key(st) in u:
            return "QUASE-INATIVO" if st == "QUASEINATIVO" else st
    return None


# ----------------------------------------------------------------------------
# Carga
# ----------------------------------------------------------------------------
def main() -> int:
    _SCRIPT_DIR  = Path(__file__).resolve().parent
    _REPO_ROOT   = _SCRIPT_DIR.parent          # rvm-designacoes-unified/
    _WORKSPACE   = _REPO_ROOT.parent           # Antigravity - RVM Designações/
    _DEFAULT_ODS = str(_WORKSPACE / "docs" / "RM Desacoplado" / "9fe36d.Relatório Mensal v03 (New).ods")

    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--ods",
        default=_DEFAULT_ODS,
        help="Workbook .ods multi-aba com mestres Glide (padrão: %(default)s)",
    )
    ap.add_argument("--dir", default=None, help="(fallback) Pasta com CSVs Glide se --ods não disponível")
    _DEFAULT_REPORTS = (
        r"C:\Users\Eliez\OneDrive\Area de Trabalho\PIONEIROS ESPECIAIS"
        r"\Estância\Sup Serviço\Relatórios Glide.xlsx"
    )
    ap.add_argument(
        "--reports",
        default=_DEFAULT_REPORTS,
        help="Relatórios Glide.xlsx p/ monthly_reports "
             f"(padrão: {_DEFAULT_REPORTS})",
    )
    ap.add_argument("--dry-run", action="store_true", help="Não grava; só reporta contagens")
    args = ap.parse_args()

    ods_path = Path(args.ods)
    dburl = os.environ.get("RM_DATABASE_URL")
    if not dburl:
        return _fail("Env RM_DATABASE_URL ausente (connection string Postgres do Supabase).")

    # -- Carrega mestres: ODS preferido; fallback para CSVs se ODS ausente
    if ods_path.exists():
        print(f"[ODS] {ods_path.name}")
        congs  = read_ods_sheet(ods_path, "Congregação")
        groups = read_ods_sheet(ods_path, "Grupos")
        pubs   = read_ods_sheet(ods_path, "PublicadorReal")
        print(f"  → {len(congs)} congregações, {len(groups)} grupos, {len(pubs)} publicadores")
    else:
        if not args.dir:
            return _fail(f"ODS não encontrado ({ods_path}) e --dir não informado.")
        base = Path(args.dir)
        def find(*names: str) -> Path:
            for n in names:
                p = base / n
                if p.exists():
                    return p
            raise FileNotFoundError(f"Nenhuma das CSVs encontrada: {names}")
        print(f"[CSV fallback] {base}")
        congs  = read_csv_rows(find("Congregacao.csv", "Congregação.csv"))
        groups = read_csv_rows(find("Grupos.csv"))
        pubs   = read_csv_rows(find("Publicador Real.csv"))
        print(f"  → {len(congs)} congregações, {len(groups)} grupos, {len(pubs)} publicadores")

    if args.dry_run:
        print("[dry-run] nada será gravado.")
        return 0

    conn = psycopg2.connect(dburl)
    conn.autocommit = False
    try:
        cur = conn.cursor()
        # Adia checagem das FKs DEFERRABLE (líderes) até o commit
        cur.execute("SET CONSTRAINTS ALL DEFERRED;")

        cong_map: dict[str, str] = {}   # glide cong id -> uuid
        group_map: dict[str, str] = {}  # glide group id -> uuid
        pub_map: dict[str, str] = {}    # glide pub id -> uuid

        # 1) Congregações
        for r in congs:
            gid = col(r, "id_Congregação", "id_Congregacao")
            name = col(r, "Nome")
            number = col(r, "Número", "Numero")
            if not name:
                continue
            cur.execute(
                "INSERT INTO rm.congregations (name, number, is_active) VALUES (%s,%s,true) RETURNING id",
                (name, number),
            )
            uid = cur.fetchone()[0]
            if gid:
                cong_map[gid] = uid

        # 2) Grupos (sem líderes ainda; guardamos glide ids de líder/ajudante)
        for r in groups:
            gid = col(r, "id_Grupo")
            cong_gid = col(r, "fk_id_Congregação", "fk_id_Congregacao")
            cong_uuid = cong_map.get(cong_gid) if cong_gid else None
            if not cong_uuid:
                continue
            number = col(r, "Número", "Numero") or "0"
            try:
                gnum = int(str(number).strip())
            except ValueError:
                gnum = 0
            cur.execute(
                """INSERT INTO rm.field_groups
                   (congregation_id, group_number, name, glide_leader_id, glide_assistant_id, is_active)
                   VALUES (%s,%s,%s,%s,%s,true)
                   ON CONFLICT (congregation_id, group_number) DO UPDATE SET name = EXCLUDED.name
                   RETURNING id""",
                (cong_uuid, gnum, col(r, "Nome do Grupo", "Nome"),
                 col(r, "id_SuperDeGrupo"), col(r, "id_SuperAJDeGrupo")),
            )
            uid = cur.fetchone()[0]
            if gid:
                group_map[gid] = uid

        # 3) Publicadores
        for r in pubs:
            gid = col(r, "id_Publicador")
            name = col(r, "Nome Completo", "NomeCompleto", "Nome")
            if not name:
                continue
            cong_uuid = cong_map.get(col(r, "fk_id_Congregação", "fk_id_Congregacao") or "")
            group_uuid = group_map.get(col(r, "id_Grupo") or "")
            cur.execute(
                """INSERT INTO rm.publishers
                   (glide_id, congregation_id, current_group_id, name, funcao, gender,
                    birth_date, field_service_status, is_active)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s,true)
                   ON CONFLICT (glide_id) DO UPDATE SET name = EXCLUDED.name
                   RETURNING id""",
                (gid, cong_uuid, group_uuid, name,
                 col(r, "Função", "Funcao", "Privilégio", "Privilegio"),
                 gender(col(r, "Sexo")),
                 to_date(col(r, "Data de Nascimento", "DataNascimento")),
                 status_norm(col(r, "Status do Último Relatório", "Status"))),
            )
            uid = cur.fetchone()[0]
            if gid:
                pub_map[gid] = uid

        # 4) Resolver líderes de grupo (2ª passada) via glide id -> pub uuid
        cur.execute("SELECT id, glide_leader_id, glide_assistant_id FROM rm.field_groups")
        for group_id, gl, ga in cur.fetchall():
            leader = pub_map.get(gl) if gl else None
            assistant = pub_map.get(ga) if ga else None
            if leader or assistant:
                cur.execute(
                    "UPDATE rm.field_groups SET leader_id = %s, assistant_leader_id = %s WHERE id = %s",
                    (leader, assistant, group_id),
                )

        # 5) sync_map (linhas base; auto-match acontece na UI Portal Sync)
        cur.execute(
            """INSERT INTO rm.publisher_sync_map (rm_publisher_id, match_status)
               SELECT id, 'unmatched' FROM rm.publishers
               ON CONFLICT (rm_publisher_id) DO NOTHING"""
        )

        # 6) monthly_reports (usa planilha canônica OneDrive por padrão)
        reports_n = 0
        _reports_path = Path(args.reports)
        if _reports_path.exists():
            reports_n = _import_reports(cur, _reports_path, cong_map, group_map, pub_map)
        else:
            print(f"[AVISO] Planilha de relatórios não encontrada: {_reports_path} — pulando monthly_reports.")

        conn.commit()
        print(f"OK: {len(cong_map)} congregações, {len(group_map)} grupos, "
              f"{len(pub_map)} publicadores, {reports_n} relatórios.")
        return 0
    except Exception as e:  # noqa: BLE001
        conn.rollback()
        return _fail(f"Rollback — erro na carga: {e}")
    finally:
        conn.close()


def _import_reports(cur, xlsx: Path, cong_map, group_map, pub_map) -> int:
    """Importa rm.monthly_reports da planilha denormalizada 'Relatórios Glide.xlsx'.
    Colunas conforme decisoes_migracao_glide_2026-06-30.json (A..BA)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("Falta openpyxl p/ ler xlsx: pip install openpyxl")
    if not xlsx.exists():
        raise FileNotFoundError(f"Planilha de relatórios não encontrada: {xlsx}")

    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(min_row=2, values_only=True)

    def cel(row: tuple[Any, ...], letter: str):
        idx = 0
        for ch in letter:
            idx = idx * 26 + (ord(ch.upper()) - 64)
        idx -= 1
        return row[idx] if idx < len(row) else None

    # Maps de fallback: publicador -> congregação, e congregação -> nome.
    # Col A/B da planilha (congregação "quando relatou") vêm nulas em muitas linhas;
    # nesse caso herdamos a congregação atual do próprio publicador.
    cur.execute("SELECT id, congregation_id FROM rm.publishers")
    pub_cong = {r[0]: r[1] for r in cur.fetchall()}
    cur.execute("SELECT id, name FROM rm.congregations")
    cong_name = {r[0]: r[1] for r in cur.fetchall()}

    n = 0
    for row in rows:
        glide_row_id = cel(row, "AC")
        pub_gid = cel(row, "AF")
        pub_uuid = pub_map.get(str(pub_gid)) if pub_gid else None
        if not pub_uuid or not glide_row_id:
            continue
        year = cel(row, "I")
        month = cel(row, "M")
        if year is None or month is None:
            continue
        glide_cong = cel(row, "A")
        # (1) congregação: A-resolvido; senão herda a do publicador
        cong_uuid = cong_map.get(str(glide_cong)) if glide_cong else None
        if cong_uuid is None:
            cong_uuid = pub_cong.get(pub_uuid)
        # congregation_at_time: snapshot B; se nulo, usa o nome da congregação resolvida
        cong_at_time = cel(row, "B") or (cong_name.get(cong_uuid) if cong_uuid else None)
        # (2) submitted_at: coluna F (data original); COALESCE p/ não violar NOT NULL
        submitted_at = cel(row, "F")
        group_name = cel(row, "N")
        cur.execute(
            """INSERT INTO rm.monthly_reports
               (publisher_id, congregation_id, congregation_at_time, group_at_time,
                reference_year, reference_month, service_year, has_preached, hours,
                bible_studies, modalities, notes, is_late_report, late_consolidation_period,
                is_auxiliary_pioneer, glide_row_id, glide_congregation_id, submitted_at)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s, COALESCE(%s, now()))
               ON CONFLICT (glide_row_id) DO NOTHING""",
            (pub_uuid, cong_uuid, cong_at_time, group_name,
             int(year), int(month), _int_or_none(cel(row, "AE")),
             _to_bool(cel(row, "D")), _num_or_none(cel(row, "R")),
             _int_or_none(cel(row, "Z")) or 0, _modalities(cel(row, "S")),
             cel(row, "AA"), _to_bool(cel(row, "K")), cel(row, "L"),
             _to_bool(cel(row, "BA")), str(glide_row_id),
             (str(glide_cong) if glide_cong else None), submitted_at),
        )
        n += 1
    return n


def _to_bool(v) -> bool:
    return str(v).strip().lower() in ("true", "sim", "1", "t", "yes") if v is not None else False


def _int_or_none(v):
    try:
        return int(v) if v not in (None, "") else None
    except (ValueError, TypeError):
        return None


def _num_or_none(v):
    try:
        return float(v) if v not in (None, "") else None
    except (ValueError, TypeError):
        return None


def _modalities(v) -> list[str]:
    if not v:
        return []
    return [m.strip() for m in str(v).replace(";", ",").split(",") if m.strip()]


def _fail(msg: str) -> int:
    print(f"ERRO: {msg}", file=sys.stderr)
    return 1


if __name__ == "__main__":
    raise SystemExit(main())

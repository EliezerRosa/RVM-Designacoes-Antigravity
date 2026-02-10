"""
Script para carregar dados do Excel para a tabela workbook_parts no Supabase.
Usa lógica de UPSERT: se parte já existe (por weekId + seq + funcao), atualiza os dados.
"""
import openpyxl
import requests
import json
from pathlib import Path
from datetime import datetime
import uuid

import os
from dotenv import load_dotenv

# Load environment variables from root .env
# Adjust path to find the .env file in the project root
root_path = Path(__file__).parent.parent.parent.parent / ".env"
load_dotenv(root_path)

# Supabase credentials (mesmas do migrate_publishers_to_supabase.py)
SUPABASE_URL = os.getenv("SUPABASE_URL_DATA")
SUPABASE_KEY = os.getenv("SUPABASE_KEY_DATA")

if not SUPABASE_URL or not SUPABASE_KEY:
    print("❌ Erro: Credenciais do Supabase não encontradas no arquivo .env")
    print(f"   Procurando em: {root_path}")
    exit(1)

def read_excel_parts(excel_path: str) -> list[dict]:
    """Lê partes do Excel gerado pelo extract_detailed_parts.py."""
    print(f"📖 Lendo Excel: {excel_path}")
    
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    # Ler cabeçalhos
    headers = [cell.value for cell in ws[1]]
    print(f"   Headers: {headers}")
    
    # Ler dados
    parts = []
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0]:  # Pular linhas vazias (checa se ID está vazio)
            continue
            
        record = dict(zip(headers, row))
        parts.append(record)
    
    print(f"   Total: {len(parts)} registros")
    return parts


def map_to_db_columns(part: dict) -> dict:
    """Mapeia nomes das colunas Excel para colunas do banco (snake_case)."""
    # Gerar batch_id único para esta importação
    batch_id = f"excel-upload-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
    
    return {
        'id': part.get('id') or str(uuid.uuid4()),
        'batch_id': batch_id,
        'week_id': part.get('weekId', ''),
        'week_display': part.get('weekDisplay', ''),
        'date': part.get('date', ''),
        'section': part.get('section', ''),
        'tipo_parte': part.get('tipoParte', ''),
        'modalidade': part.get('modalidade', ''),
        'titulo_parte': part.get('tituloParte', ''),
        'descricao_parte': part.get('descricaoParte', ''),
        'detalhes_parte': part.get('detalhesParte', ''),
        'seq': part.get('seq', 0),
        'funcao': part.get('funcao', 'Titular'),
        'duracao': str(part.get('duracao', '')),
        'hora_inicio': part.get('horaInicio', ''),
        'hora_fim': part.get('horaFim', ''),
        'raw_publisher_name': part.get('rawPublisherName', ''),
        'status': part.get('status', 'PENDING'),
        'created_at': datetime.now().isoformat(),
    }


def upsert_to_supabase(parts: list[dict]) -> tuple[int, int]:
    """
    Envia partes para Supabase via REST API com UPSERT.
    Retorna (sucesso, total).
    """
    url = f"{SUPABASE_URL}/rest/v1/workbook_parts"
    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
        # UPSERT: se conflito no ID, atualiza a linha existente
        "Prefer": "resolution=merge-duplicates"
    }
    
    # Mapear para formato de banco de dados
    db_rows = [map_to_db_columns(p) for p in parts]
    
    # Usar o mesmo batch_id para todos
    batch_id = db_rows[0]['batch_id'] if db_rows else f"batch-{datetime.now().isoformat()}"
    for row in db_rows:
        row['batch_id'] = batch_id
    
    print(f"\n📤 Enviando {len(db_rows)} registros para Supabase...")
    print(f"   Batch ID: {batch_id}")
    
    # Dividir em chunks de 100 para evitar timeout
    chunk_size = 100
    success_count = 0
    
    for i in range(0, len(db_rows), chunk_size):
        chunk = db_rows[i:i+chunk_size]
        response = requests.post(url, headers=headers, json=chunk)
        
        if response.status_code in (200, 201):
            success_count += len(chunk)
            print(f"   ✅ Chunk {i//chunk_size + 1}: {len(chunk)} registros")
        else:
            print(f"   ❌ Erro no chunk {i//chunk_size + 1}: {response.status_code}")
            print(f"      {response.text[:500]}")
    
    return success_count, len(db_rows)


def main():
    # Caminho do Excel gerado pelo script offline (na pasta pai do projeto)
    excel_path = Path(__file__).parent.parent.parent / "dados_sensiveis" / "ANTIGRAVITY Apostilas" / "partes_v14.xlsx"
    
    if not excel_path.exists():
        print(f"❌ Arquivo não encontrado: {excel_path}")
        return
    
    # Ler dados do Excel
    parts = read_excel_parts(str(excel_path))
    
    if not parts:
        print("❌ Nenhum registro encontrado no Excel")
        return
    
    # Enviar para Supabase
    success, total = upsert_to_supabase(parts)
    
    print(f"\n{'='*60}")
    if success == total:
        print(f"✅ Upload completo! {success}/{total} registros enviados.")
    else:
        print(f"⚠️ Upload parcial: {success}/{total} registros enviados.")
    print('='*60)


if __name__ == "__main__":
    main()
